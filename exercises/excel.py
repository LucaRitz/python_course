# Access Exel documents
from openpyxl import Workbook, load_workbook

# To sleep a few second between emails
import time

# Import smtplib for the actual sending function
import smtplib

# Import the email modules we'll need
from email.message import EmailMessage


def send(username: str, file: str) -> None:

    print(username + ", ", end="")
    password = input("enter your password: ")

    # set up the SMTP server
    with smtplib.SMTP(host='mail.bfh.ch', port=587) as s:
        s.starttls()
        s.login(username, password)

        wb: Workbook = load_workbook(file)
        sheet = wb.active

        for person in __read_person(sheet):
            s.sendmail('ritz.luca@gmail.com', person.mail, person.message())


class Person:
    def __init__(self, name: str, firstname: str, mark: str, mail: str):
        self.name = name
        self.firstname = firstname
        self.mark = mark
        self.mail = mail

    def message(self):
        return 'Hey %s %s, your mark is %s!' % (self.firstname, self.name, self.mark)


def __read_person(sheet) -> iter:
    for i in range(4, sheet.max_row + 1):
        lastname: str = sheet.cell(row = i, column = 2).value
        if lastname != '' and lastname is not None:
            yield Person(lastname, sheet.cell(row=i, column=3).value, sheet.cell(row=i, column=10).value,
                         sheet.cell(row=i, column=11).value)
